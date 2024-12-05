import openpyxl
import matplotlib.pyplot as plt
from telegram import Update
from telegram.ext import Application, CommandHandler, ContextTypes
from flask import Flask, request
import os

# رابط السيرفر على Render مع المسار
NGROK_URL = "https://telegram-student-bot-li3t.onrender.com/webhook"  # استبدل بالرابط الفعلي من Render

# إعداد Flask
app = Flask(__name__)

# قراءة بيانات الطالب من ملف Excel بناءً على ورقته
def get_student_grades(student_id):
    # تحميل ملف Excel باستخدام المسار الكامل
    file_path = "C:/Users/NOUR-ALSHAM/OneDrive/Desktop/my bot/students_grades.xlsx"  # استبدل هذا بالمسار الفعلي لملفك
    wb = openpyxl.load_workbook(file_path)
    
    # التحقق إذا كانت ورقة الطالب موجودة
    if str(student_id) in wb.sheetnames:
        sheet = wb[str(student_id)]
    else:
        return None  # ورقة الطالب غير موجودة
    
    # استخراج البيانات من الورقة
    student_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        student_data.append(row)
    
    return student_data

# عرض الدرجات مع علامات أوراق العمل والرسم البياني
async def show_grades(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    student_id = context.args[0]  # رقم الطالب من المدخلات
    
    # الحصول على درجات الطالب
    student_data = get_student_grades(student_id)
    
    if student_data:
        # حساب المعدل العام للدرجات
        grades = [grade[1] for grade in student_data]  # درجات الامتحانات
        homework_marks = [grade[2] for grade in student_data]  # درجات أوراق العمل
        average_grade = sum(grades) / len(grades)
        
        # رسم بياني للدرجات
        subjects = [grade[0] for grade in student_data]
        
        # إنشاء الرسم البياني
        fig, ax = plt.subplots(figsize=(8, 6))
        ax.bar(subjects, grades, color='skyblue', label='Grades')
        ax.bar(subjects, homework_marks, color='lightgreen', label='Homework Marks', alpha=0.6)
        
        ax.set_xlabel('Subjects')
        ax.set_ylabel('Marks')
        ax.set_title(f"Grades and Homework Marks for Student {student_id}")
        ax.legend()
        plt.tight_layout()
        plt.savefig('grades_and_homework_chart.png')
        
        # إرسال الرسالة مع الرسم البياني
        await update.message.reply_text(f"معدل الطالب {student_id}: {average_grade:.2f}")
        await update.message.reply_photo(photo=open('grades_and_homework_chart.png', 'rb'))
        
        # إرسال رابط السيرفر مع النتائج
        await update.message.reply_text(f"يمكنك الوصول إلى رابط الخدمة عبر السيرفر من هنا: {NGROK_URL}")
    else:
        await update.message.reply_text(f"لم يتم العثور على درجات للطالب {student_id}.")

# إعداد البوت
def main():
    application = Application.builder().token("8048197051:AAHvIPxu_OtD5G66CgVmJBai4mjFOqDw-cc").build()  # استبدل بـ التوكن الخاص بك
    application.add_handler(CommandHandler("grades", show_grades))  # عرض الدرجات مع علامات أوراق العمل

    # إعداد webhook
    application.bot.set_webhook(url=NGROK_URL)  # رابط السيرفر + المسار

    # تشغيل البوت باستخدام webhook
    application.run_webhook(listen="0.0.0.0", port=5000, url_path="webhook")  # المسار هنا هو "webhook"

# إعداد Flask لاستقبال التحديثات
@app.route("/webhook", methods=["POST"])  # المسار هنا هو "webhook"
def webhook():
    json_str = request.get_data().decode("UTF-8")
    update = Update.de_json(json_str, application.bot)
    application.process_update(update)
    return "OK"

if __name__ == '__main__':
    # تشغيل Flask
    app.run(debug=True, host='0.0.0.0', port=5000)
